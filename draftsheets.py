"""Re-score a FantasyPros DraftSheets workbook for this league and merge its values into the tool.

The workbook (Scoring / QB / RB / WR / TE / ECR / RISK / Rookies sheets) computes everything with
Excel formulas driven by the ``Scoring`` inputs.  Without Excel we cannot recalculate it, so this
module re-implements that math in Python: per-position points from the raw stat projections,
the games-missed discount, the ECR-blended "Zscore Projection", positional baselines (including
the FLEX split), VBD and tiers.  ``validate()`` checks the replication against the workbook's
cached values under its own settings; ``compute()`` then runs it under ours.

Usage::

    python draftsheets.py "~/Downloads/DraftSheets_2026 (Copy).xlsx"            # re-score + merge
    python draftsheets.py "<xlsx>" --validate                                     # check math only

Outputs ``data/draftsheets_2026.csv`` (the re-scored sheet) and adds ``VBD``, ``ExternalTier``,
``RankAvg`` (ECR overall rank) and ``DraftSheetPts`` columns to ``data/players.csv``.
"""
from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from pathlib import Path

from models import load_config, normalize_name, normalize_team, settings_from_config

POSITIONS = ("QB", "RB", "WR", "TE")
# Aggregate table sizes per position (rows the workbook actually ranks) and the smaller range its
# tier / slot-points lookups use.
TABLE_ROWS = {"QB": 48, "RB": 100, "WR": 100, "TE": 100}
LOOKUP_ROWS = {"QB": 50, "RB": 100, "WR": 100, "TE": 50}
TIER_GAP = {"QB": 0.2, "RB": 0.15, "WR": 0.15, "TE": 0.2}
TIER_CAP = {"QB": 15, "TE": 15}
# FLEX sheet candidates: ranks (starters+1 .. starters+20) at RB/WR, (+1 .. +12) at TE.
FLEX_CANDIDATES = {"RB": 20, "WR": 20, "TE": 12}


# --------------------------------------------------------------------------- #
# Reading the workbook
# --------------------------------------------------------------------------- #

def _num(v) -> float:
    try:
        return float(v) if v not in (None, "", "-") else 0.0
    except (TypeError, ValueError):
        return 0.0


def read_workbook(path) -> dict:
    """Pull the raw inputs out of the workbook: stat projections, ECR, risk table, rookies, settings."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    raw = {"stats": {}, "ecr": [], "risk": {}, "rookies": set()}

    for pos in POSITIONS:
        ws = wb[pos]
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        table = {}
        i = 0
        while i < len(rows):
            r = rows[i]
            name = r[0]
            if name and str(name).strip() and str(name).strip() != "\xa0":
                avg = r
                high = rows[i + 1] if i + 1 < len(rows) else ()
                low = rows[i + 2] if i + 2 < len(rows) else ()
                table[str(name).strip()] = {
                    "team": r[1], "avg": avg, "high": high, "low": low, "header": header,
                }
                i += 3
            else:
                i += 1
        raw["stats"][pos] = table

    ws = wb["ECR"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[2] or not r[4]:
            continue
        m = re.match(r"([A-Z]+)(\d+)", str(r[4]))
        if not m or m.group(1) not in POSITIONS:
            continue
        raw["ecr"].append({
            "rank": int(_num(r[0])), "tier": int(_num(r[1])), "name": str(r[2]).strip(), "team": r[3],
            "pos": m.group(1), "pos_rank": int(m.group(2)), "bye": r[5],
            "upside": r[6], "bust": r[7], "sos": r[8], "ecr_vs_adp": r[9],
        })

    for r in wb["RISK"].iter_rows(min_row=2, max_col=2, values_only=True):
        if r[0] and r[1] is not None:
            raw["risk"][str(r[0])] = float(r[1])

    for r in wb["Rookies"].iter_rows(min_row=2, max_col=1, values_only=True):
        if r[0]:
            raw["rookies"].add(str(r[0]).strip())

    s = wb["Scoring"]
    g = lambda ref: s[ref].value  # noqa: E731
    raw["settings"] = {
        "teams": int(g("I4")), "QB": int(g("J4")), "RB": int(g("K4")), "WR": int(g("L4")),
        "TE": int(g("M4")), "FLEX": int(g("N4")), "BENCH": int(g("O4")), "SUPERFLEX": int(g("P4")),
        "PassINCMP": _num(g("B5")), "PassCMP": _num(g("B6")), "PassYDS": _num(g("B7")),
        "PassTDs": _num(g("B8")), "INTS": _num(g("B9")), "SACKS": _num(g("B10")),
        "RushATT": _num(g("B12")), "RushYDS": _num(g("B13")), "RushTDS": _num(g("B14")),
        "PPR_RB": _num(g("B16")), "PPR_WR": _num(g("B17")), "PPR_TE": _num(g("B18")),
        "RecYDS": _num(g("B19")), "RecTDS": _num(g("B20")), "FL": _num(g("B22")),
        "Pass1D": _num(g("B23")), "Rush1D": _num(g("B24")), "Rec1D": _num(g("B25")),
        "flex_positions": ("RB", "WR", "TE"),
    }
    # Cached outputs, for validation.
    ag = wb["Aggregate"]
    cached = {}
    col0 = {"QB": 1, "RB": 15, "WR": 29, "TE": 43}   # A, O, AC, AQ
    for pos, c in col0.items():
        for r in ag.iter_rows(min_row=3, max_row=2 + TABLE_ROWS[pos], min_col=c, max_col=c + 12, values_only=True):
            if r[1]:
                cached[(pos, str(r[1]).strip())] = {"zscore": r[8], "vbd": r[9], "tier": r[11], "missed": r[12]}
    raw["cached"] = cached
    wb.close()
    return raw


def settings_from_league(cfg: dict, base: dict) -> dict:
    """Translate config.yaml (roster + Yahoo stat-id scoring) into the workbook's Scoring inputs."""
    st = settings_from_config(cfg)
    sc = st.scoring or {}
    s = dict(base)
    s.update({
        "teams": st.num_teams,
        "QB": st.roster.starters("QB"), "RB": st.roster.starters("RB"),
        "WR": st.roster.starters("WR"), "TE": st.roster.starters("TE"),
        "FLEX": st.roster.flex_slots, "BENCH": st.roster.bench_slots, "SUPERFLEX": 0,
        "flex_positions": tuple(st.roster.flex_positions),
    })
    if sc:
        per_pt = lambda v: (1.0 / v) if v else 0.0  # noqa: E731
        s.update({
            "PassYDS": per_pt(float(sc.get("4", 0.04))), "PassTDs": float(sc.get("5", 4)),
            "INTS": float(sc.get("6", -1)), "RushYDS": per_pt(float(sc.get("9", 0.1))),
            "RushTDS": float(sc.get("10", 6)), "RecYDS": per_pt(float(sc.get("12", 0.1))),
            "RecTDS": float(sc.get("13", 6)), "FL": float(sc.get("18", -2)),
        })
        ppr = float(sc.get("11", 0.0))
        s.update({"PPR_RB": ppr, "PPR_WR": ppr, "PPR_TE": ppr})
    return s


# --------------------------------------------------------------------------- #
# The workbook's math
# --------------------------------------------------------------------------- #

def _points(pos: str, row: tuple, s: dict) -> float:
    """Raw season points for one stat row (the workbook's per-position POINTS formula, no rookie add)."""
    if not row:
        return 0.0
    if pos == "QB":
        att, cmp_, yds, tds, ints, ratt, ryds, rtds, fl = (_num(x) for x in row[2:11])
        pa1d = ((yds / cmp_) / 10.96) * 0.5218 * cmp_ if cmp_ else 0.0
        ru1d = (((ryds / ratt) / 5.3) * 0.401) * ratt if ratt else 0.0
        core = (cmp_ * s["PassCMP"] + yds / s["PassYDS"] + tds * s["PassTDs"] + ints * s["INTS"]
                + ratt * s["RushATT"] + ryds / s["RushYDS"] + rtds * s["RushTDS"] + fl * s["FL"]
                + pa1d * s["Pass1D"] + ru1d * s["Rush1D"])
        return max(core, 0.0) + att * s["SACKS"] * 0.07 + (att - cmp_) * s["PassINCMP"]
    if pos == "RB":
        att, yds, tds, rec, ryds, rtds, fl = (_num(x) for x in row[2:9])
        ru1d = (yds / att) / 4.5 * 0.25 * att if att else 0.0
        rec1d = (ryds / rec) / 7.95 * 0.376 * rec if rec else 0.0
        return (att * s["RushATT"] + yds / s["RushYDS"] + tds * s["RushTDS"] + rec * s["PPR_RB"]
                + ryds / s["RecYDS"] + rtds * s["RecTDS"] + fl * s["FL"] + ru1d * s["Rush1D"] + rec1d * s["Rec1D"])
    if pos == "WR":
        rec, yds, tds, att, ryds, rtds, fl = (_num(x) for x in row[2:9])
        ru1d = ryds / att / 6.89 * 0.427 * att if att else 0.0
        rec1d = yds / rec / 12.87 * 0.633 * rec if rec else 0.0
        return (rec * s["PPR_WR"] + yds / s["RecYDS"] + tds * s["RecTDS"] + att * s["RushATT"]
                + ryds / s["RushYDS"] + rtds * s["RushTDS"] + fl * s["FL"] + ru1d * s["Rush1D"] + rec1d * s["Rec1D"])
    rec, yds, tds, fl = (_num(x) for x in row[2:6])
    rec1d = yds / rec / 11.1 * 0.585 * rec if rec else 0.0
    return rec * s["PPR_TE"] + yds / s["RecYDS"] + tds * s["RecTDS"] + fl * s["FL"] + rec1d * s["Rec1D"]


def _rookie_bonus(pos: str, avg_pts: float, is_rookie: bool) -> float:
    if not is_rookie:
        return 0.0
    per_game = avg_pts / 17
    if pos == "RB":
        return 17 * max(0.0, 0.258 * (14.9 - per_game))
    if pos == "WR":
        return 17 * max(0.0, 0.28 * (12 - per_game))
    return 0.0


def _large(values: list, k: int):
    vals = sorted((v for v in values if v is not None), reverse=True)
    return vals[k - 1] if 0 < k <= len(vals) else None


def _tiers(vbds: list, gap: float, cap: int | None) -> dict:
    """Workbook tier table: walk VBD descending; new tier when the drop from the tier's first VBD
    exceeds ``gap`` x (2nd-largest VBD).  Returns {vbd_value: tier}."""
    ranked = sorted(vbds, reverse=True)
    if not ranked:
        return {}
    second = ranked[1] if len(ranked) > 1 else ranked[0]
    out, tier, tier_top = {}, 1, ranked[0]
    for v in ranked:
        if (tier_top - v) > gap * second:
            tier, tier_top = tier + 1, v
        if cap:
            tier = min(tier, cap)
        out.setdefault(v, tier)
    return out


def compute(raw: dict, s: dict) -> list:
    """Run the workbook's pipeline under scoring/roster settings ``s``.  Returns row dicts."""
    stats, ecr, risk, rookies = raw["stats"], raw["ecr"], raw["risk"], raw["rookies"]
    tables = {pos: [] for pos in POSITIONS}
    for e in ecr:
        pos = e["pos"]
        if e["pos_rank"] > TABLE_ROWS[pos]:
            continue
        missed = risk.get(f"{pos}{e['pos_rank']}")
        if missed is None:  # beyond the RISK table: use the deepest value for that position
            missed = max((v for k, v in risk.items() if k.startswith(pos)), default=0.0)
        st = stats[pos].get(e["name"])
        low = avg = high = 0.0
        if st:
            bonus = _rookie_bonus(pos, _points(pos, st["avg"], s), e["name"] in rookies)
            scale = (16 - missed) / 17
            avg = (_points(pos, st["avg"], s) + bonus) * scale
            high = (_points(pos, st["high"], s) + bonus) * scale
            low = (_points(pos, st["low"], s) + bonus) * scale
        tables[pos].append({**e, "low": low, "avg": avg, "high": high, "missed": missed})

    # ECR-blended projection: average of low/avg/high and the AVG of whoever is k-th best at the position.
    for pos, rows in tables.items():
        avgs = [r["avg"] for r in rows[:LOOKUP_ROWS[pos]]]
        for r in rows:
            slot = _large(avgs, r["pos_rank"])
            parts = [r["low"], r["avg"], r["high"]] + ([slot] if slot is not None else [])
            r["zscore"] = sum(parts) / len(parts)

    # Baselines (Scoring rows 40-43) with the FLEX split (FLEX sheet).
    teams = s["teams"]
    starters = {"QB": (s["QB"] + s["SUPERFLEX"]) * teams, "RB": round(s["RB"] * teams * 1.267),
                "WR": round(s["WR"] * teams * 1.23), "TE": round(s["TE"] * teams)}
    flex_n = round(s["FLEX"] * teams * 1.4)
    cands = []
    for pos in ("RB", "WR", "TE"):
        if pos not in s["flex_positions"]:
            continue
        avgs = [r["avg"] for r in tables[pos][:LOOKUP_ROWS[pos]]]
        for k in range(1, FLEX_CANDIDATES[pos] + 1):
            v = _large(avgs, starters[pos] + k)
            if v is not None:
                cands.append((v, pos))
    cands.sort(reverse=True)
    flex_share = {"RB": 0, "WR": 0, "TE": 0}
    for _, pos in cands[:flex_n]:
        flex_share[pos] += 1
    baseline_rank = {"QB": round(starters["QB"] * 1.17), "RB": starters["RB"] + flex_share["RB"],
                     "WR": starters["WR"] + flex_share["WR"], "TE": starters["TE"] + flex_share["TE"]}
    z = {pos: [r["zscore"] for r in tables[pos]] for pos in POSITIONS}
    base_pts = {"RB": _large(z["RB"], baseline_rank["RB"]) or 0.0, "WR": _large(z["WR"], baseline_rank["WR"]) or 0.0}
    base_pts["TE"] = base_pts["WR"] if baseline_rank["TE"] == 0 else (_large(z["TE"], baseline_rank["TE"]) or 0.0)
    base_pts["QB"] = max(_large(z["QB"][:LOOKUP_ROWS["QB"]], baseline_rank["QB"]) or 0.0,
                         base_pts["RB"], base_pts["WR"], base_pts["TE"])

    out = []
    for pos, rows in tables.items():
        for r in rows:
            r["vbd"] = r["zscore"] - base_pts[pos]
        tiers = _tiers([r["vbd"] for r in rows[:LOOKUP_ROWS[pos]]], TIER_GAP[pos], TIER_CAP.get(pos))
        last_tier = max(tiers.values()) if tiers else 1
        for r in rows:
            r["sheet_tier"] = tiers.get(r["vbd"], last_tier)
            out.append(r)
    out.sort(key=lambda r: r["rank"])
    meta = {"starters": starters, "flex_n": flex_n, "flex_share": flex_share,
            "baseline_rank": baseline_rank, "baseline_pts": base_pts}
    return out, meta


# --------------------------------------------------------------------------- #
# Validation, output, merge
# --------------------------------------------------------------------------- #

def validate(raw: dict, tol: float = 0.05) -> dict:
    """Recompute under the workbook's own settings and diff against its cached Aggregate values."""
    rows, meta = compute(raw, raw["settings"])
    cached = raw["cached"]
    mism, n = [], 0
    for r in rows:
        c = cached.get((r["pos"], r["name"]))
        if not c or c["zscore"] is None:
            continue
        n += 1
        dz = abs(r["zscore"] - _num(c["zscore"]))
        dv = abs(r["vbd"] - _num(c["vbd"]))
        dt = (r["sheet_tier"] != c["tier"]) if c["tier"] not in (None, "") else False
        if dz > tol or dv > tol or dt:
            mism.append((r["pos"], r["name"], round(dz, 3), round(dv, 3), r["sheet_tier"], c["tier"]))
    return {"compared": n, "mismatches": mism, "meta": meta}


OUT_FIELDS = ["Player", "Position", "Team", "Bye", "ECRRank", "ECRTier", "PosRank", "Upside", "Bust",
              "MissedGames", "Low", "Avg", "High", "Pts", "VBD", "Tier", "ECRvsADP"]


def write_sheet_csv(rows: list, path, note: str) -> None:
    with open(path, "w", newline="") as f:
        f.write(f"# {note}\n")
        w = csv.writer(f)
        w.writerow(OUT_FIELDS)
        for r in rows:
            w.writerow([r["name"], r["pos"], normalize_team(r["team"]), r["bye"] if r["bye"] not in (None, "-") else "",
                        r["rank"], r["tier"], r["pos_rank"], r["upside"], r["bust"], round(r["missed"], 3),
                        round(r["low"], 2), round(r["avg"], 2), round(r["high"], 2), round(r["zscore"], 2),
                        round(r["vbd"], 2), r["sheet_tier"], r["ecr_vs_adp"] if r["ecr_vs_adp"] != "-" else ""])


def merge_into_players(rows: list, players_csv) -> dict:
    """Add VBD / ExternalTier / RankAvg / DraftSheetPts columns to data/players.csv, matching on
    normalised name + position (team as tiebreak).  Other columns and the comment line are kept."""
    path = Path(players_csv)
    lines = path.read_text().splitlines()
    comments = [ln for ln in lines if ln.startswith("#")]
    body = [ln for ln in lines if not ln.startswith("#")]
    reader = csv.DictReader(body)
    fields = list(reader.fieldnames or [])
    for col in ("VBD", "ExternalTier", "RankAvg", "DraftSheetPts"):
        if col not in fields:
            fields.append(col)
    by_key = {}
    for r in rows:
        by_key.setdefault((normalize_name(r["name"]), r["pos"]), []).append(r)
    matched, used = 0, set()
    out_rows = []
    for m in reader:
        cands = by_key.get((normalize_name(m["Player"]), m["Position"]), [])
        if len(cands) > 1:
            same = [c for c in cands if normalize_team(c["team"]) == normalize_team(m["Team"])]
            cands = same or cands
        if cands:
            r = cands[0]
            used.add(id(r))
            matched += 1
            m["VBD"] = round(r["vbd"], 2)
            m["ExternalTier"] = r["sheet_tier"]
            m["RankAvg"] = r["rank"]
            m["DraftSheetPts"] = round(r["zscore"], 2)
        else:
            for col in ("VBD", "ExternalTier", "RankAvg", "DraftSheetPts"):
                m.setdefault(col, "")
        out_rows.append(m)
    unmatched = [r for r in rows if id(r) not in used]
    with path.open("w", newline="") as f:
        for c in comments:
            f.write(c + "\n")
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(out_rows)
    return {"matched": matched, "unmatched": unmatched}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx")
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--out", default="data/draftsheets_2026.csv")
    ap.add_argument("--players", default=None, help="master table to enrich (default: paths.players_csv)")
    ap.add_argument("--validate", action="store_true", help="only check the replication against cached values")
    ap.add_argument("--no-merge", action="store_true")
    a = ap.parse_args(argv)

    raw = read_workbook(Path(a.xlsx).expanduser())
    v = validate(raw)
    print(f"validation vs workbook cache: {v['compared']} players compared, {len(v['mismatches'])} mismatches")
    for m in v["mismatches"][:15]:
        print("   ", m)
    if a.validate:
        return 1 if v["mismatches"] else 0

    cfg = load_config(a.config)
    s = settings_from_league(cfg, raw["settings"])
    rows, meta = compute(raw, s)
    print("league settings:", {k: s[k] for k in ("teams", "QB", "RB", "WR", "TE", "FLEX", "BENCH", "PPR_RB", "PassYDS", "INTS", "FL", "flex_positions")})
    print("baselines:", meta)
    note = (f"DraftSheets re-scored for {cfg['league'].get('name')} ({s['teams']} tm, QB{s['QB']}/RB{s['RB']}/WR{s['WR']}/TE{s['TE']}"
            f"/FLEX{s['FLEX']} {'/'.join(s['flex_positions'])}, {s['PPR_RB']:g} PPR). Pts = ECR-blended 16-game projection; VBD vs baselines "
            + ", ".join(f"{p}{meta['baseline_rank'][p]}" for p in POSITIONS) + ".")
    write_sheet_csv(rows, a.out, note)
    print(f"wrote {a.out}: {len(rows)} players")
    if not a.no_merge:
        players_csv = a.players or cfg["paths"]["players_csv"]
        res = merge_into_players(rows, players_csv)
        top_unmatched = [f"{r['name']} ({r['pos']}{r['pos_rank']})" for r in res["unmatched"] if r["rank"] <= 150]
        print(f"merged into {players_csv}: {res['matched']} matched; unmatched in ECR top 150: {top_unmatched or 'none'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
